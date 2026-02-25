"""
Tool: read_file_header

Reads the first N rows and column names from tabular data files
(CSV, TSV, Excel, fixed-width, SAS, Stata, Parquet).

Optimized for token efficiency — avoids loading full files when only
peeking at structure. Uses format-specific metadata APIs for row counts.
"""

from __future__ import annotations

import os

import pandas as pd

from versifai.core.tools.base import BaseTool, ToolResult

# Encodings to try in order when reading CSV-like files
_ENCODING_FALLBACKS = ["utf-8", "latin-1", "cp1252", "iso-8859-1"]


class FileReaderTool(BaseTool):
    @property
    def name(self) -> str:
        return "read_file_header"

    @property
    def description(self) -> str:
        return (
            "Read the column names and first N sample rows from a tabular data file "
            "(CSV, TSV, Excel, Parquet, fixed-width, SAS, Stata). "
            "Returns column names, inferred dtypes, sample values, and row count estimate. "
            "Use this to understand the structure and content of a data file."
        )

    @property
    def parameters_schema(self) -> dict:
        return {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Absolute path to the data file.",
                },
                "n_rows": {
                    "type": "integer",
                    "description": "Number of sample rows to return. Default 10. Use small values (10-15) for peeking at structure.",
                    "default": 10,
                },
                "encoding": {
                    "type": "string",
                    "description": (
                        "Character encoding to use (e.g. utf-8, latin-1). "
                        "If omitted, the tool auto-detects by trying common encodings."
                    ),
                },
                "separator": {
                    "type": "string",
                    "description": "Column delimiter for CSV/TSV files. Auto-detected if omitted.",
                },
                "sheet_name": {
                    "type": "string",
                    "description": "Sheet name or index for Excel files. Defaults to first sheet.",
                },
                "skip_rows": {
                    "type": "integer",
                    "description": "Number of rows to skip at the top of the file. Default 0.",
                    "default": 0,
                },
            },
            "required": ["file_path"],
        }

    def _execute(
        self,
        file_path: str,
        n_rows: int = 10,
        encoding: str | None = None,
        separator: str | None = None,
        sheet_name: str | None = None,
        skip_rows: int = 0,
        **kwargs,
    ) -> ToolResult:
        if not os.path.isfile(file_path):
            return ToolResult(success=False, error=f"File not found: {file_path}")

        ext = os.path.splitext(file_path)[1].lower().lstrip(".")

        if ext in ("csv", "tsv", "txt", "dat"):
            return self._read_csv(file_path, n_rows, encoding, separator, skip_rows)
        elif ext in ("xlsx", "xls"):
            return self._read_excel(file_path, n_rows, sheet_name, skip_rows)
        elif ext == "parquet":
            return self._read_parquet(file_path, n_rows)
        elif ext == "sas7bdat":
            return self._read_sas(file_path, n_rows)
        elif ext == "dta":
            return self._read_stata(file_path, n_rows)
        else:
            return ToolResult(
                success=False,
                error=f"Unsupported file extension: .{ext}. Supported: csv, tsv, txt, dat, xlsx, xls, parquet, sas7bdat, dta",
            )

    # ------------------------------------------------------------------
    # Format-specific readers
    # ------------------------------------------------------------------

    def _read_csv(
        self,
        file_path: str,
        n_rows: int,
        encoding: str | None,
        separator: str | None,
        skip_rows: int,
    ) -> ToolResult:
        encodings_to_try = [encoding] if encoding else _ENCODING_FALLBACKS
        last_error = ""

        for enc in encodings_to_try:
            try:
                # First, try to detect the separator if not provided
                sep = separator
                if not sep:
                    with open(file_path, encoding=enc) as f:
                        first_line = f.readline()
                    if "\t" in first_line:
                        sep = "\t"
                    elif "|" in first_line:
                        sep = "|"
                    else:
                        sep = ","

                # Read a sample to get column names and dtypes
                df_sample = pd.read_csv(
                    file_path,
                    sep=sep,
                    encoding=enc,
                    nrows=n_rows,
                    skiprows=range(1, skip_rows + 1) if skip_rows else None,
                    on_bad_lines="warn",
                    low_memory=False,
                )

                # Estimate total row count from file size ratio
                # using the sample we already read (no extra reads)
                file_size = os.path.getsize(file_path)
                sample_rows = len(df_sample)
                if sample_rows > 0:
                    # Estimate bytes-per-row from the sample we already have
                    # Use the sample's memory footprint as a proxy
                    sample_bytes = df_sample.memory_usage(deep=True).sum()
                    if sample_bytes > 0 and sample_rows >= n_rows:
                        # File has more rows than we sampled — estimate
                        bytes_per_row = file_size / sample_rows  # rough
                        estimated_rows = (
                            int(file_size / bytes_per_row) if bytes_per_row > 0 else sample_rows
                        )
                    else:
                        estimated_rows = sample_rows
                else:
                    estimated_rows = 0

                return self._build_result(df_sample, file_path, enc, estimated_rows)

            except (UnicodeDecodeError, UnicodeError) as e:
                last_error = f"Encoding {enc} failed: {e}"
                continue
            except Exception as e:
                last_error = str(e)
                continue

        return ToolResult(
            success=False, error=f"Failed to read CSV with all encodings. Last error: {last_error}"
        )

    def _read_excel(
        self, file_path: str, n_rows: int, sheet_name: str | None, skip_rows: int
    ) -> ToolResult:
        sheet = sheet_name if sheet_name else 0

        # Get sheet names and row count without loading full data
        xls = pd.ExcelFile(file_path)
        all_sheets = xls.sheet_names

        df_sample = pd.read_excel(file_path, sheet_name=sheet, nrows=n_rows, skiprows=skip_rows)

        # Estimate row count from openpyxl metadata (avoids loading full file)
        estimated_rows = -1
        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".xlsx":
            try:
                from openpyxl import load_workbook

                wb = load_workbook(file_path, read_only=True, data_only=True)
                target_sheet = sheet if isinstance(sheet, str) else all_sheets[sheet]
                ws = wb[target_sheet]
                estimated_rows = ws.max_row - 1 - skip_rows  # subtract header + skipped
                wb.close()
            except Exception:
                estimated_rows = -1  # fallback: unknown
        if estimated_rows < 0:
            # For .xls or if openpyxl fails, use file size heuristic
            estimated_rows = max(len(df_sample), int(os.path.getsize(file_path) / 200))

        result = self._build_result(df_sample, file_path, "excel", estimated_rows)
        result.data["sheets"] = all_sheets
        return result

    def _read_parquet(self, file_path: str, n_rows: int) -> ToolResult:
        import pyarrow.parquet as pq

        # Get row count from parquet metadata (zero data reads)
        pf = pq.ParquetFile(file_path)
        total_rows = pf.metadata.num_rows

        # Read only the first n_rows using pyarrow (efficient)
        table = pf.read_row_groups([0]) if pf.metadata.num_row_groups > 0 else pf.read()
        df_sample = table.to_pandas().head(n_rows)

        return self._build_result(df_sample, file_path, "parquet", total_rows)

    def _read_sas(self, file_path: str, n_rows: int) -> ToolResult:
        # Use iterator to read only n_rows (avoids loading full file)
        reader = pd.read_sas(file_path, encoding="latin-1", chunksize=n_rows)
        df_sample = next(reader)
        reader.close()

        # SAS files store row count in header metadata
        try:
            reader2 = pd.read_sas(file_path, encoding="latin-1", chunksize=1)
            # The reader object has row_count from the SAS header
            estimated_rows = getattr(reader2, "row_count", -1)
            reader2.close()
            if estimated_rows <= 0:
                estimated_rows = int(os.path.getsize(file_path) / 200)  # rough estimate
        except Exception:
            estimated_rows = int(os.path.getsize(file_path) / 200)

        return self._build_result(df_sample, file_path, "sas", estimated_rows)

    def _read_stata(self, file_path: str, n_rows: int) -> ToolResult:
        # Use iterator to read only n_rows (avoids loading full file)
        reader = pd.read_stata(file_path, chunksize=n_rows)
        df_sample = next(reader)
        reader.close()

        # Stata header contains observation count — extract via StataReader
        try:
            with pd.io.stata.StataReader(file_path) as sr:
                estimated_rows = sr.nobs
        except Exception:
            estimated_rows = int(os.path.getsize(file_path) / 200)

        return self._build_result(df_sample, file_path, "stata", estimated_rows)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _build_result(
        self, df: pd.DataFrame, file_path: str, encoding: str, estimated_rows: int
    ) -> ToolResult:
        columns_info = []
        for col in df.columns:
            columns_info.append(
                {
                    "name": str(col),
                    "dtype": str(df[col].dtype),
                    "sample_values": [str(v) for v in df[col].head(3).tolist()],
                    "null_count_in_sample": int(df[col].isna().sum()),
                }
            )

        # Only include a few sample rows — enough to see the shape, not flood tokens
        sample_data = df.head(5).fillna("").astype(str).to_dict(orient="records")

        return ToolResult(
            success=True,
            data={
                "file_path": file_path,
                "encoding": encoding,
                "column_count": len(df.columns),
                "columns": [str(c) for c in df.columns],
                "columns_detail": columns_info,
                "estimated_total_rows": estimated_rows,
                "sample_row_count": len(df),
                "sample_data": sample_data,
                "file_size_mb": round(os.path.getsize(file_path) / (1024 * 1024), 2),
            },
            summary=(
                f"Read {len(df)} sample rows from {os.path.basename(file_path)} "
                f"({len(df.columns)} columns, ~{estimated_rows:,} total rows, encoding={encoding})"
            ),
        )
